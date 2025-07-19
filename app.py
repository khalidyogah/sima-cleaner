import streamlit as st
import calendar
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
import io

def clean_excel(file, key_column='A'):
    wb = load_workbook(file)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for sheet in wb.worksheets:
        for mrange in list(sheet.merged_cells):
            sheet.unmerge_cells(str(mrange))

        sheet.delete_rows(1, 6)

        col_index = ord(key_column.upper()) - ord('A') + 1
        last_data_row = 0
        for row in reversed(range(1, sheet.max_row + 1)):
            if sheet.cell(row=row, column=col_index).value not in (None, ""):
                last_data_row = row
                break
        if last_data_row < sheet.max_row:
            sheet.delete_rows(last_data_row + 1, sheet.max_row - last_data_row)

        for col in reversed(range(1, sheet.max_column + 1)):
            if all(sheet.cell(row, col).value in [None, ""] for row in range(1, sheet.max_row + 1)):
                sheet.delete_cols(col)

        for row in reversed(range(1, sheet.max_row + 1)):
            if all(sheet.cell(row, col).value in [None, ""] for col in range(1, sheet.max_column + 1)):
                sheet.delete_rows(row)

        for col in sheet.columns:
            sheet.column_dimensions[get_column_letter(col[0].column)].width = 15
        for rdim in sheet.row_dimensions.values():
            rdim.height = 15

        for row in range(2, sheet.max_row + 1):
            cell = sheet[f"F{row}"]
            val = cell.value
            if isinstance(val, str):
                try:
                    parsed = datetime.strptime(val.strip(), "%b-%y")
                    last_day = calendar.monthrange(parsed.year, parsed.month)[1]
                    cell.value = datetime(parsed.year, parsed.month, last_day)
                    cell.number_format = "DD/MM/YYYY"
                except:
                    pass
            elif isinstance(val, datetime):
                cell.number_format = "DD/MM/YYYY"

        for row in range(2, sheet.max_row + 1):
            for col in range(ord("I"), ord("O")+1):
                cell = sheet.cell(row=row, column=col - ord("A") + 1)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'

        sheet._images.clear()
        sheet.freeze_panes = sheet["A3"]

        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.border = thin_border

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

# Streamlit UI
st.title("Excel Cleaner Tool")
uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx"])

if uploaded_file:
    filename = uploaded_file.name.replace('.xlsx', '') + '-clean.xlsx'
    cleaned_bytes = clean_excel(uploaded_file)

    st.success("✅ Cleaning done!")
    st.download_button("Download Cleaned Excel", data=cleaned_bytes, file_name=filename)
